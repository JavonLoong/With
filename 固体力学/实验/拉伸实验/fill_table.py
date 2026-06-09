"""
从原始拉伸 CSV 自动提取三种材料关键载荷，并回填 Excel 第 2 节“实验数据”。

默认数据源：
- 低碳钢: jwl.csv (回退到 jjs.csv)
- 铸铁:   jwl0.csv (回退到 jjs0.csv)
- 高分子: AB44高分子材料拉伸.csv
"""

import csv
import io
import math
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


def read_csv_rows(filename):
    raw = Path(filename).read_bytes()
    text = None
    for enc in ('utf-8-sig', 'gb18030', 'gbk', 'latin1'):
        try:
            text = raw.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        raise ValueError(f'无法解码文件: {filename}')
    return list(csv.reader(io.StringIO(text)))


def try_float(x):
    try:
        return float(str(x).strip().replace('\ufeff', ''))
    except Exception:
        return None


def extract_load_columns(rows):
    """根据单位行识别所有载荷列（单位 N，排除 N/mm2）。"""
    if len(rows) < 3:
        return []
    units = rows[2]
    cols = []
    for i, u in enumerate(units):
        uu = str(u).strip().upper()
        if uu == 'N':
            cols.append(i)
    return cols


def extract_series_from_col(rows, col):
    vals = []
    for r in rows[3:]:
        if col < len(r):
            v = try_float(r[col])
            if v is not None:
                vals.append(v)
    return vals


def get_single_load_series(filename):
    rows = read_csv_rows(filename)
    load_cols = extract_load_columns(rows)
    if not load_cols:
        raise ValueError(f'{filename} 中未识别到载荷列(单位 N)')
    return extract_series_from_col(rows, load_cols[0])


def get_multi_specimen_maxloads(filename):
    rows = read_csv_rows(filename)
    load_cols = extract_load_columns(rows)
    if not load_cols:
        raise ValueError(f'{filename} 中未识别到载荷列(单位 N)')
    peaks = []
    for col in load_cols:
        series = extract_series_from_col(rows, col)
        if series:
            peaks.append(max(series))
    return peaks


def sig3_kn(force_n):
    """N -> kN 并保留 3 位有效数字。"""
    k_n = force_n / 1000.0
    if k_n == 0:
        return 0.0
    mag = math.floor(math.log10(abs(k_n)))
    factor = 10 ** (mag - 2)
    return round(k_n / factor) * factor


def pick_first_existing(candidates):
    for f in candidates:
        if Path(f).exists():
            return f
    return None


def find_steel_loads():
    steel_file = pick_first_existing(['jwl.csv', 'jjs.csv'])
    if steel_file is None:
        raise FileNotFoundError('未找到低碳钢数据文件: jwl.csv/jjs.csv')
    loads = get_single_load_series(steel_file)
    if not loads:
        raise ValueError(f'{steel_file} 没有可用载荷数据')

    n = len(loads)
    front_n = max(1000, min(8000, int(n * 0.35)))
    sub = loads[:front_n]
    feh = max(sub)
    idx = sub.index(feh)

    # 上屈服点后短窗口内寻找下屈服极小值
    win = max(300, int(n * 0.10))
    plateau = loads[idx:min(n, idx + win)]
    fel = min(plateau) if plateau else feh

    fm = max(loads)
    return steel_file, feh, fel, fm


def find_cast_fm():
    cast_file = pick_first_existing(['jwl0.csv', 'jjs0.csv'])
    if cast_file is None:
        raise FileNotFoundError('未找到铸铁数据文件: jwl0.csv/jjs0.csv')
    peaks = get_multi_specimen_maxloads(cast_file)
    if not peaks:
        raise ValueError(f'{cast_file} 没有可用载荷数据')
    return cast_file, sum(peaks) / len(peaks), peaks


def find_polymer_fm():
    poly_file = pick_first_existing(['AB44高分子材料拉伸.csv'])
    if poly_file is None:
        raise FileNotFoundError('未找到高分子数据文件: AB44高分子材料拉伸.csv')
    loads = get_single_load_series(poly_file)
    if not loads:
        raise ValueError(f'{poly_file} 没有可用载荷数据')
    return poly_file, max(loads)


def set_merged(ws, row, col, value, center, fill):
    c = ws.cell(row=row, column=col)
    c.value = value
    c.alignment = center
    c.fill = fill


def upsert_note(ws, text):
    marker = '数据来源（自动填充）:'
    target_row = None
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if isinstance(v, str) and v.startswith(marker):
            target_row = r
            break

    if target_row is None:
        target_row = ws.max_row + 1
        ws.merge_cells(start_row=target_row, start_column=1, end_row=target_row, end_column=16)

    note_cell = ws.cell(row=target_row, column=1)
    note_cell.value = marker + text
    note_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    note_cell.font = Font(size=10, color='555555')
    ws.row_dimensions[target_row].height = 36


def main():
    xlsx = Path('拉伸实验数据记录表.xlsx')
    if not xlsx.exists():
        raise FileNotFoundError('未找到 拉伸实验数据记录表.xlsx，请先运行 create_table.py 生成模板')

    steel_file, feh_n, fel_n, fm_s_n = find_steel_loads()
    cast_file, fm_ci_n, cast_peaks_n = find_cast_fm()
    poly_file, fm_p_n = find_polymer_fm()

    feh_kn = sig3_kn(feh_n)
    fel_kn = sig3_kn(fel_n)
    fms_kn = sig3_kn(fm_s_n)
    fmci_kn = sig3_kn(fm_ci_n)
    fmp_kn = sig3_kn(fm_p_n)

    print('【自动提取结果（kN，3 位有效数字）】')
    print(f'  低碳钢 数据源: {steel_file}')
    print(f'    F_eH={feh_kn}, F_eL={fel_kn}, Fm={fms_kn}')
    print(f'  铸铁 数据源: {cast_file}')
    print(f'    试样峰值N={", ".join(f"{v:.1f}" for v in cast_peaks_n)}')
    print(f'    Fm(均值)={fmci_kn}')
    print(f'  高分子 数据源: {poly_file}')
    print(f'    Fm={fmp_kn}')

    wb = load_workbook(xlsx)
    ws = wb.active

    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    fill_y = PatternFill(fill_type='solid', fgColor='FFFDE7')

    # Section 2: 行 13/14/15, 列 4/8/12
    set_merged(ws, 13, 4, feh_kn, center, fill_y)
    set_merged(ws, 13, 8, fel_kn, center, fill_y)
    set_merged(ws, 13, 12, fms_kn, center, fill_y)

    set_merged(ws, 14, 12, fmci_kn, center, fill_y)
    set_merged(ws, 15, 12, fmp_kn, center, fill_y)

    upsert_note(
        ws,
        (
            f'低碳钢={steel_file} (F_eH={feh_kn} kN, F_eL={fel_kn} kN, Fm={fms_kn} kN); '
            f'铸铁={cast_file} (Fm均值={fmci_kn} kN); '
            f'高分子={poly_file} (Fm={fmp_kn} kN)。'
            '本脚本仅自动填写第2节可由曲线直接提取的载荷项。'
        ),
    )

    wb.save(xlsx)
    print(f'\n✓ 已写入 {xlsx}')


if __name__ == '__main__':
    main()
