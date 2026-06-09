import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "拉伸实验数据"

# ---- 样式 ----
thin = Side(style='thin')
def tb():
    return Border(left=thin, right=thin, top=thin, bottom=thin)

center = Alignment(horizontal='center', vertical='center', wrap_text=True)
left   = Alignment(horizontal='left',   vertical='center', wrap_text=True)
bold   = Font(bold=True, size=11)
normal = Font(size=11)

def mc(sr, sc, er, ec, value, align=None, font=None):
    """合并单元格并写入内容"""
    if (sr, sc) != (er, ec):
        ws.merge_cells(start_row=sr, start_column=sc, end_row=er, end_column=ec)
    cell = ws.cell(row=sr, column=sc)
    cell.value = value
    cell.alignment = align or center
    cell.font = font or normal
    return cell

def apply_border(sr, sc, er, ec):
    for r in range(sr, er + 1):
        for c in range(sc, ec + 1):
            ws.cell(row=r, column=c).border = tb()

COLS = 16   # 总列数 A~P

# ============================================================
# 标题
# ============================================================
mc(1, 1, 1, COLS,
   '实验数据记录应包括低碳钢（或铝合金）、铸铁、高分子三种材料。',
   align=center, font=bold)
ws.row_dimensions[1].height = 28

# ============================================================
# 第 1 节：试件尺寸
# 列布局：
#   A(1)=材料名称  B(2)=L₀mm
#   C(3)~E(5)=截面1①②平均  F(6)~H(8)=截面2①②平均  I(9)~K(11)=截面3①②平均
#   L(12)=d₀mm  M(13)=S₀mm²
#   N(14)=Lumm  O(15)=dumm  P(16)=Summ²
# ============================================================
mc(2, 1, 2, COLS, '1. 试件尺寸', font=bold)
ws.row_dimensions[2].height = 20

# -- 第 3 行：最顶层标题 --
mc(3, 1, 6, 1,  '材料\n名称')           # A3:A6
mc(3, 2, 6, 2,  'L₀\n(mm)')            # B3:B6
mc(3, 3, 3, 11, '试    验    前')       # C3:K3  (d 测量列)
mc(3, 12, 6, 12,'d₀\nmm')              # L3:L6
mc(3, 13, 6, 13,'S₀\nmm²')             # M3:M6
mc(3, 14, 3, 16,'试 验 后')             # N3:P3

# -- 第 4 行 --
mc(4, 3, 4, 11, 'd（mm）')              # C4:K4
mc(4, 14, 6, 14,'L u\nmm')             # N4:N6
mc(4, 15, 6, 15,'d u\nmm')             # O4:O6
mc(4, 16, 6, 16,'S u\nmm²')            # P4:P6

# -- 第 5 行 --
mc(5, 3, 5, 5,  '截  面  1')           # C5:E5
mc(5, 6, 5, 8,  '截  面  2')           # F5:H5
mc(5, 9, 5, 11, '截  面  3')           # I5:K5

# -- 第 6 行：①②平均 --
for offset, label in enumerate(['①','②','平均','①','②','平均','①','②','平均']):
    c = ws.cell(row=6, column=3+offset)
    c.value = label
    c.alignment = center
    c.font = normal

apply_border(3, 1, 6, COLS)
for r in range(3, 7):
    ws.row_dimensions[r].height = 20

# -- 数据行 --
# 低碳钢/铝合金
mc(7, 1, 7, 1, '低碳钢\n/铝合金')
apply_border(7, 1, 7, COLS)
ws.row_dimensions[7].height = 28

# 铸铁
mc(8, 1, 8, 1, '铸铁')
mc(8, 2, 8, 11,
   '直径 d₀ =          mm，    夹具间距 L =          mm', align=left)
apply_border(8, 1, 8, COLS)
ws.row_dimensions[8].height = 28

# 高分子
mc(9, 1, 9, 1, '高分子')
mc(9, 2, 9, 11,
   '宽度 b =          mm，    高度 h =          mm，    夹具间距 L =          mm', align=left)
apply_border(9, 1, 9, COLS)
ws.row_dimensions[9].height = 28

# ============================================================
# 第 2 节：实验数据
# 列布局（共享同 16 列）：
#   A(1:3)=材料  D(4:7)=FeH  H(8:11)=FeL  L(12:16)=Fm
# ============================================================
ws.row_dimensions[10].height = 8

mc(10, 1, 10, COLS, '2. 实验数据', font=bold)
ws.row_dimensions[10].height = 20

mc(11, 1, 12, 3,  '材  料')
mc(11, 4, 12, 7,  '上屈服载荷\nF_eH（KN）')
mc(11, 8, 12, 11, '（下）屈服载荷\nF_eL（或 F_P0.2）（KN）')
mc(11, 12, 12, 16,'最大载荷\nF_m（KN）')
apply_border(11, 1, 12, COLS)
ws.row_dimensions[11].height = 24
ws.row_dimensions[12].height = 24

# 低碳钢/铝合金
mc(13, 1, 13, 3, '低碳钢\n/铝合金')
apply_border(13, 1, 13, COLS)
ws.row_dimensions[13].height = 28

# 铸铁
mc(14, 1, 14, 3, '铸铁')
mc(14, 4, 14, 7, '/')
mc(14, 8, 14, 11,'/')
apply_border(14, 1, 14, COLS)
ws.row_dimensions[14].height = 28

# 高分子
mc(15, 1, 15, 3, '高分子')
mc(15, 4, 15, 7, '/')
apply_border(15, 1, 15, COLS)
ws.row_dimensions[15].height = 28

# ============================================================
# 第 3 节：计算结果
# 列布局：
#   A(1:2)=材料  C(3:4)=E  E(5:7)=ReH  H(8:10)=ReL  K(11:12)=Rm  M(13:14)=A  O(15:16)=Z
# ============================================================
ws.row_dimensions[16].height = 8

mc(16, 1, 16, COLS, '3. 计算结果', font=bold)
ws.row_dimensions[16].height = 20

# 双行表头
mc(17, 1, 18, 2,  '材  料')
mc(17, 3, 17, 4,  '刚度指标')
mc(17, 5, 17, 12, '强    度    指    标')
mc(17, 13, 17, 16,'塑  性  指  标')

mc(18, 3, 18, 4,  '弹性模量\nE（GPa）')
mc(18, 5, 18, 7,  '上屈服强度\nR_eH（MPa）')
mc(18, 8, 18, 10, '屈服强度/下屈服强度\nR_eL（R_P0.2）（MPa）')
mc(18, 11, 18, 12,'抗拉强度\nR_m（MPa）')
mc(18, 13, 18, 14,'断后伸长率\nA（%）')
mc(18, 15, 18, 16,'断面收缩率\nZ（%）')
apply_border(17, 1, 18, COLS)
ws.row_dimensions[17].height = 22
ws.row_dimensions[18].height = 34

# 低碳钢/铝合金（全部空白，待填）
mc(19, 1, 19, 2, '低碳钢\n/铝合金')
apply_border(19, 1, 19, COLS)
ws.row_dimensions[19].height = 28

# 铸铁（无屈服点，无塑性变形）
mc(20, 1, 20, 2, '铸铁')
mc(20, 3, 20, 4, '/')       # E
mc(20, 5, 20, 7, '/')       # ReH
mc(20, 8, 20, 10,'/')       # ReL
# Rm：留空，待计算填写
mc(20, 13, 20, 14,'/')      # A
mc(20, 15, 20, 16,'/')      # Z
apply_border(20, 1, 20, COLS)
ws.row_dimensions[20].height = 28

# 高分子（无上/下屈服强度，无断面收缩率）
mc(21, 1, 21, 2, '高分子')
mc(21, 5, 21, 7, '/')       # ReH
mc(21, 8, 21, 10,'/')       # ReL
# E、Rm、A：留空，待填
mc(21, 15, 21, 16,'/')      # Z
apply_border(21, 1, 21, COLS)
ws.row_dimensions[21].height = 28

# ============================================================
# 注释
# ============================================================
ws.row_dimensions[22].height = 8
mc(23, 1, 23, COLS,
   '注：1. 实验结果采用三位有效数字。', align=left)
mc(24, 1, 24, COLS,
   '    2. 如断后伸长率采用短比例试样的标距，则用 A 表示；否则将实际原始标距长度'
   '作为 A 的下标。如 L₀=60mm，A₆₀ₘₘ', align=left)
ws.row_dimensions[23].height = 20
ws.row_dimensions[24].height = 28

# ============================================================
# 列宽
# ============================================================
ws.column_dimensions['A'].width = 8
ws.column_dimensions['B'].width = 7
for col_idx in range(3, 12):          # C ~ K
    ws.column_dimensions[get_column_letter(col_idx)].width = 5
ws.column_dimensions['L'].width = 7
ws.column_dimensions['M'].width = 8
ws.column_dimensions['N'].width = 7
ws.column_dimensions['O'].width = 7
ws.column_dimensions['P'].width = 8

# 打印区域 & 页面设置
ws.print_area = f'A1:P24'
ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
ws.page_setup.fitToPage = True
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 0

output = r'd:\虚拟C盘\实验\拉伸实验数据记录表.xlsx'
wb.save(output)
print(f'已保存：{output}')
